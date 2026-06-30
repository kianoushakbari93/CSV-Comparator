#!/usr/bin/env python3
"""
CSV Comparator - DuckDB Edition
High-Performance Migration Validation (Hive to Snowflake)

DuckDB replaces pandas as the core data engine:
  - Streams CSVs from disk (no 12GB+ RAM for two 6GB files)
  - SQL-based normalisation, hashing, and matching
  - Only pulls small unmatched sets into pandas for fuzzy comparison
  - Report generation unchanged (pandas + openpyxl for XLSX)

All features from the original are preserved:
  - Intelligent delimiter detection & CSV preprocessing
  - Value normalisation (timestamps, booleans, nulls, numerics)
  - Composite key auto-detection
  - Fuzzy key matching
  - Duplicate detection (opt-in via --detect-duplicates)
  - Parallel XLSX report generation
  - Batch folder mode
"""

from __future__ import annotations

import pandas as pd
import duckdb
import sys
import re
import os
import csv
import tempfile
import argparse
import logging
import multiprocessing as mp
from datetime import datetime
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor, as_completed
from contextlib import contextmanager
from typing import Any, Optional, Dict, List, Tuple, Set, DefaultDict


class CSVComparatorError(Exception):
    """Raised when the CSV comparator encounters a fatal error."""


# =============================================================================
# LOGGING
# =============================================================================

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

# =============================================================================
# CONSTANTS
# =============================================================================

COLUMN_EXCLUSION_PATTERNS: List[str] = [
    r'.*_?DESC.*', r'.*_?DESCRIPTION.*', r'.*_?COMMENT.*', r'.*_?NOTE.*', r'.*_?TEXT.*',
    r'.*_?AMOUNT.*', r'.*_?VALUE.*', r'.*_?PRICE.*', r'.*_?QUANTITY.*', r'.*_?QTY.*',
    r'.*_?BALANCE.*', r'.*_?TOTAL.*', r'.*_?PERCENT.*', r'.*_?PCT.*', r'.*_?RATIO.*', r'.*_?RATE.*',
    r'CREATED.*', r'UPDATED.*', r'MODIFIED.*',
    r'.*_?TIMESTAMP.*', r'.*_?DATETIME.*', r'.*_?TIME$', r'.*_?DT$', r'.*_?TS$',
    r'.*_?FILE.*', r'.*_?PATH.*', r'.*_?FILENAME.*',
    r'.*_?LOAD.*', r'.*_?ETL.*', r'.*_?BATCH.*', r'.*_?RUN.*', r'.*_?PROCESS.*',
    r'.*_?AUDIT.*', r'.*_?INSERT.*', r'.*_?UPDATE.*',
    r'^ROW_.*', r'.*_?SOURCE$', r'.*_?SRC$', r'.*_?NAME$', r'^NAME_?.*',
    r'.*_?STATUS.*', r'.*_?FLAG.*', r'.*_?IND$', r'.*_?INDICATOR.*',
    r'.*_?UUID.*', r'.*_?GUID.*', r'.*_?HASH$',
    r'^EDD_.*', r'^ETL_.*', r'^DW_.*', r'^DWH_.*',
    r'.*_?EXTRACT_.*', r'^EXTRACT_.*', r'^GENERATED_.*', r'.*_GENERATED$',
]

# Column name patterns that are strong composite key candidates.
# Columns matching these get a priority boost and are NOT excluded even
# if their values look like dates/timestamps (e.g. AS_OF_DATE).
KEY_CANDIDATE_PATTERNS: List[str] = [
    # ID/code columns — almost always good keys
    r'.*_?ID$', r'^ID_.*',                      # SEC_ID, PORTFOLIO_ID, ID_TYPE
    r'.*_?CD$', r'.*_?CODE$',                   # SERVICE_CD, FUND_CD, ISIN_CD, ACCOUNT_CODE
    r'.*_?KEY$', r'.*_?NUM$', r'.*_?NUMBER$',   # TRADE_KEY, ACCOUNT_NUM, ORDER_NUMBER
    r'.*_?REF$', r'.*_?REFERENCE$',             # TRADE_REF, ORDER_REFERENCE
    # Common business date keys — dates used as partitioning/grouping keys
    r'^AS_OF_DATE$', r'.*_?AS_OF_DATE$',        # AS_OF_DATE, REPORT_AS_OF_DATE
    r'^TRADE_DATE$', r'.*_?TRADE_DATE$',        # TRADE_DATE
    r'^EFFECTIVE_DATE$', r'.*_?EFFECTIVE_DATE$', # EFFECTIVE_DATE
    r'^SETTLE_DATE$', r'.*_?SETTLE_DATE$',      # SETTLE_DATE, SETTLEMENT_DATE
    r'^BUSINESS_DATE$', r'.*_?BUSINESS_DATE$',  # BUSINESS_DATE
    r'^VALUATION_DATE$', r'.*_?VALUATION_DATE$',# VALUATION_DATE
    r'^REPORT_DATE$', r'.*_?REPORT_DATE$',      # REPORT_DATE
    r'^MONTHENDDATE$', r'^MONTH_END_DATE$',     # MONTHENDDATE
    # Identifiers commonly used as composite keys in financial data
    r'^ISIN$', r'^CUSIP$', r'^SEDOL$', r'^TICKER$', r'^SYMBOL$',
    r'^ACCOUNT$', r'^FUND$', r'^PORTFOLIO$', r'^SECURITY$',
]

DEFAULT_TIMESTAMP_PRECISION: Optional[int] = None
DEFAULT_NUMERIC_PRECISION: int = 6

NULL_LIKE_VALUES: Set[str] = {
    '', 'null', 'none', 'nan', 'nat', 'n/a', 'na', '#n/a', '<null>', '"<null>"', '\\n',
}


# =============================================================================
# PYTHON HELPERS (kept for fuzzy matching on small unmatched sets)
# =============================================================================

def _is_null_like(value: Any) -> bool:
    if pd.isna(value) or value is None:
        return True
    return str(value).strip().strip('"').lower() in NULL_LIKE_VALUES


def normalise_value(
    value: Any, skip_normalisation: bool = False,
    timestamp_precision: Optional[int] = DEFAULT_TIMESTAMP_PRECISION,
    numeric_precision: int = DEFAULT_NUMERIC_PRECISION
) -> Optional[str]:
    """Python normalise_value - used for fuzzy matching on small sets.
    Bulk normalisation uses SQL (_build_normalise_sql)."""
    if pd.isna(value):
        return None
    str_value = str(value)
    if skip_normalisation:
        return None if str_value.lower().strip() in ['', 'nan', 'nat'] else str_value
    if str_value.lower().strip() in ['', 'null', 'none', 'nan', 'nat', 'n/a', 'na', '#n/a', '<null>', '\\n']:
        return None
    stripped = str_value.strip()
    if stripped.lower() in ['true', 'yes', 'y', '1']:
        return 'true'
    if stripped.lower() in ['false', 'no', 'n', '0']:
        return 'false'
    # Strip zero-offset timezone suffixes (Snowflake TIMESTAMP_TZ exports):
    # 'Z', ' UTC', ' GMT', '+00:00', '+0000', '-00:00', '-0000'
    ts_candidate = re.sub(r'\s?(Z|UTC|GMT|[+-]00:?00)$', '', stripped)
    ts = re.match(r'^(\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2})\.(\d+)$', ts_candidate)
    if ts:
        dt_part = ts.group(1).replace('T', ' ')
        if timestamp_precision is None or timestamp_precision == 0:
            return dt_part
        frac = ts.group(2)[:timestamp_precision].ljust(timestamp_precision, '0')
        return f"{dt_part}.{frac}"
    if re.match(r'^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}$', ts_candidate):
        return ts_candidate.replace('T', ' ')
    iso = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', stripped)
    if iso:
        y, m, d = int(iso.group(1)), int(iso.group(2)), int(iso.group(3))
        if 1900 <= y <= 2100 and 1 <= m <= 12 and 1 <= d <= 31:
            return stripped
    try:
        cn = stripped.replace(',', '')
        # Expand scientific notation (Hive doubles export as 1.5E7) to
        # positional form so it matches Snowflake's expanded output.
        # Skipped when the positional form exceeds DECIMAL(38,18) capacity,
        # mirroring the SQL normalisation's TRY_CAST fallthrough.
        if re.match(r'^[+-]?(\d+\.?\d*|\.\d+)[eE][+-]?\d+$', cn):
            from decimal import Decimal, InvalidOperation
            try:
                d = Decimal(cn)
                positional = format(d.normalize(), 'f')
                int_digits = len(positional.split('.')[0].lstrip('+-'))
                frac_digits = len(positional.split('.')[1]) if '.' in positional else 0
                # Symmetric with SQL DECIMAL(33,18) guard: ≤15 integer digits,
                # ≤18 fractional digits (the exactly-expandable range)
                if int_digits <= 15 and frac_digits <= 18:
                    cn = positional
            except (InvalidOperation, ValueError):
                pass
        if re.match(r'^[+-]?\d+$', cn):
            return str(int(cn))
        elif re.match(r'^[+-]?\d+\.0+$', cn):
            # Use the integer part directly (exact) rather than int(float(cn)),
            # which loses precision for >15-digit values and would diverge from
            # the SQL path's exact HUGEINT cast (e.g. 12345678901234567890.0).
            return str(int(cn.split('.', 1)[0]))
        elif re.match(r'^[+-]?(\d+\.?\d*|\.\d+)$', cn):
            if '.' in cn:
                ip, dp = cn.split('.', 1)
                if len(dp) > numeric_precision:
                    dp = dp[:numeric_precision]
                dp = dp.rstrip('0')
                return f"{ip}.{dp}" if dp else ip
            return cn
    except (ValueError, TypeError):
        pass
    return stripped


def normalise_key_for_fuzzy(value: Any) -> str:
    if pd.isna(value) or value is None:
        return 'NULL'
    s = str(value).strip()
    if s.lower() in ['', 'null', 'none', 'nan', 'nat', 'n/a', 'na', '<null>']:
        return 'NULL'
    n = re.sub(r'[/|\\]', ' ', s)
    n = re.sub(r'\s+', ' ', n)
    n = re.sub(r'[_\-\.\(\)\[\]\{\}]', '', n)
    return n.strip().upper()


def key_values_fuzzy_equal(
    src_val: Any, tgt_val: Any, skip_normalisation: bool = False,
    numeric_precision: int = DEFAULT_NUMERIC_PRECISION,
    numeric_tolerance: float = 100.0, numeric_tolerance_pct: float = 0.01
) -> Tuple[bool, str]:
    src_n = normalise_value(src_val, skip_normalisation, numeric_precision=numeric_precision)
    tgt_n = normalise_value(tgt_val, skip_normalisation, numeric_precision=numeric_precision)
    if src_n == tgt_n:
        return True, 'EXACT'
    if src_n is None and tgt_n is None:
        return True, 'EXACT'
    if src_n is None or tgt_n is None:
        return False, 'DIFFERENT'
    try:
        sn = float(str(src_n).replace(',', ''))
        tn = float(str(tgt_n).replace(',', ''))
        diff = abs(sn - tn)
        mx = max(abs(sn), abs(tn), 1.0)
        if diff <= numeric_tolerance or (diff / mx) <= numeric_tolerance_pct:
            return True, 'FUZZY_NUMERIC'
        return False, 'DIFFERENT'
    except (ValueError, TypeError):
        pass
    sf, tf = normalise_key_for_fuzzy(src_n), normalise_key_for_fuzzy(tgt_n)
    if sf == tf:
        return True, 'FUZZY_STRING'
    if len(sf) > 0 and len(tf) > 0:
        md = max(2, min(len(sf), len(tf)) // 10)
        if abs(len(sf) - len(tf)) <= md:
            longer = sf if len(sf) >= len(tf) else tf
            shorter = tf if len(sf) >= len(tf) else sf
            if shorter in longer:
                return True, 'FUZZY_STRING'
    return False, 'DIFFERENT'


def build_fuzzy_key(row: dict, key_cols: List[str]) -> str:
    parts = []
    for c in key_cols[:3]:
        parts.append(f"{c}={normalise_key_for_fuzzy(row.get(c))}")
    return '||'.join(parts)


def normalised_values_equal(src_norm: Any, tgt_norm: Any,
                            skip_normalisation: bool = False,
                            numeric_precision: int = DEFAULT_NUMERIC_PRECISION) -> bool:
    """
    Equality check for two already-normalised values in deep comparison.

    String equality (case-insensitive unless skip_normalisation) first.
    For numerics, additionally allows |a-b| <= 10^-precision: when the same
    double is exported by Hive in 7-sig-fig scientific notation but by
    Snowflake in expanded decimals, truncation at the configured precision
    can wobble by exactly one unit in the last place. That wobble is a
    rendering artifact, not a data difference.
    """
    s, t = str(src_norm), str(tgt_norm)
    if s == t:
        return True
    if not skip_normalisation and s.lower() == t.lower():
        return True
    if skip_normalisation:
        return False
    try:
        sf_, tf_ = float(s), float(t)
        return abs(sf_ - tf_) <= 10.0 ** (-numeric_precision)
    except (ValueError, TypeError):
        return False


# =============================================================================
# CSV PREPROCESSING (runs before DuckDB loading)
# =============================================================================

def detect_delimiter_for_line(line: str) -> str:
    delimiters: Dict[str, int] = {'|': 0, ',': 0, '\t': 0, '~': 0, ';': 0}
    for d in delimiters:
        delimiters[d] = line.count(d)
    best = max(delimiters, key=delimiters.get)
    try:
        dialect = csv.Sniffer().sniff(line, delimiters='|,\t~;')
        if dialect.delimiter == best or delimiters[best] == 0:
            return dialect.delimiter
    except csv.Error:
        pass
    return ',' if delimiters[best] == 0 else best


def detect_delimiters(filepath: str, sample_lines: int = 10) -> Tuple[str, str]:
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = [f.readline().strip() for _ in range(sample_lines + 1) if True]
            lines = [l for l in lines if l]
        if not lines:
            return ',', ','
        hd = detect_delimiter_for_line(lines[0])
        data = lines[1:]
        if not data:
            return hd, hd
        rd = None
        hcc = len(list(csv.reader([lines[0]], delimiter=hd))[0])
        try:
            dialect = csv.Sniffer().sniff('\n'.join(data), delimiters='|,\t~;')
            counts = [len(list(csv.reader([dl], delimiter=dialect.delimiter))[0]) for dl in data]
            if all(c == hcc for c in counts):
                rd = dialect.delimiter
        except csv.Error:
            pass
        if rd is None:
            ds: Dict[str, int] = {'|': 0, ',': 0, '\t': 0, '~': 0, ';': 0}
            for line in data:
                for d in ds:
                    ds[d] += line.count(d)
            rd = max(ds, key=ds.get)
            if ds[rd] == 0:
                rd = hd
        # Validate: if the chosen row delimiter gives a different column count
        # than the header delimiter, the row delimiter is wrong (e.g. commas
        # inside pipe-delimited numbers like 1,234,567). Fall back to header.
        if rd != hd:
            try:
                row_counts = [len(list(csv.reader([dl], delimiter=rd))[0]) for dl in data]
                if not all(c == hcc for c in row_counts):
                    # Try header delimiter on data rows
                    hd_counts = [len(list(csv.reader([dl], delimiter=hd))[0]) for dl in data]
                    if all(c == hcc for c in hd_counts):
                        rd = hd
            except Exception:
                pass
        return hd, rd
    except Exception:
        return ',', ','


def _convert_double_escaped_record(record: str) -> str:
    inner = record[1:-1]
    parts = inner.split('""|""')
    parts[0] = parts[0][2:]
    parts[-1] = parts[-1][:-2]
    return '|'.join('"' + p + '"' for p in parts)


def preprocess_quoted_rows(filepath: str) -> Tuple[str, int]:
    qfix = defix = 0
    has_crlf = False
    with open(filepath, 'rb') as f:
        if b'\r\n' in f.read(65536):
            has_crlf = True

    def classify(rec: str) -> str:
        s = rec.rstrip('\r\n')
        if not (s.startswith('"') and s.endswith('"') and len(s) > 1):
            return 'ok'
        if s.startswith('"""') and s.endswith('"""') and len(s) > 5 and '""|""' in s:
            return 'double_escaped'
        if s.startswith('""') and s.endswith('""') and len(s) > 3:
            return 'errant'
        for p in ['","', '"|"', '"\t"', '";"', '"~"']:
            if p in s:
                return 'ok'
        return 'errant'

    def iter_recs(f):
        acc = None
        for line in f:
            s = line.rstrip('\r\n')
            if acc is not None:
                acc.append(s)
                if s.endswith('"""'):
                    yield '\n'.join(acc)
                    acc = None
                continue
            if s.startswith('"""') and '""|""' in s and not s.endswith('"""'):
                acc = [s]
                continue
            yield s
        if acc:
            yield '\n'.join(acc)

    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        f.readline()
        for rec in iter_recs(f):
            k = classify(rec)
            if k == 'double_escaped':
                defix += 1
            elif k == 'errant':
                qfix += 1

    total = qfix + defix
    if total == 0 and not has_crlf:
        return filepath, 0
    if defix > 0:
        logger.info(f"  [!] Detected {defix:,} double-escaped rows - collapsing to standard quoting")

    fd, tmp = tempfile.mkstemp(suffix='.csv', prefix='cleaned_')
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as inf, \
         os.fdopen(fd, 'w', encoding='utf-8', newline='') as out:
        out.write(inf.readline().rstrip('\r\n') + '\n')
        for rec in iter_recs(inf):
            k = classify(rec)
            if k == 'double_escaped':
                rec = _convert_double_escaped_record(rec)
            elif k == 'errant':
                rec = rec[1:-1]
            out.write(rec + '\n')
    return tmp, total


@contextmanager
def cleaned_csv(filepath: str):
    cp, n = preprocess_quoted_rows(filepath)
    try:
        yield cp, n
    finally:
        if cp != filepath:
            try:
                os.remove(cp)
            except OSError:
                pass


# =============================================================================
# DUCKDB SQL NORMALISATION
# =============================================================================

def _build_normalise_sql(col_expr: str, skip_normalisation: bool = False,
                         numeric_precision: int = DEFAULT_NUMERIC_PRECISION,
                         timestamp_precision: Optional[int] = DEFAULT_TIMESTAMP_PRECISION) -> str:
    """Build SQL CASE expression replicating normalise_value() logic.

    Uses TRY_CAST and LIKE patterns instead of REGEXP_MATCHES for
    dramatically better performance on wide tables (100x+ faster).
    """
    c = col_expr
    v = f'CAST({c} AS VARCHAR)'
    tv = f'TRIM({v})'
    ltv = f'LOWER({tv})'
    cn = f"REPLACE({tv},',','')"  # comma-stripped value

    if skip_normalisation:
        return f"CASE WHEN {c} IS NULL THEN NULL WHEN {ltv} IN ('','nan','nat') THEN NULL ELSE {v} END"

    prec = numeric_precision

    # Timezone-stripped candidate for timestamp checks (Snowflake TIMESTAMP_TZ
    # exports zero offsets as 'Z', ' UTC', '+00:00', '+0000', etc.)
    tsv = f"REGEXP_REPLACE({tv}, '\\s?(Z|UTC|GMT|[+-]00:?00)$', '')"

    # Timestamp fractional handling (applied to tz-stripped candidate)
    if timestamp_precision is None or timestamp_precision == 0:
        # Strip fractional: just take first 19 chars (YYYY-MM-DD HH:MM:SS)
        ts_frac_result = f"LEFT(REPLACE({tsv},'T',' '),19)"
    else:
        p = timestamp_precision
        ts_frac_result = (
            f"LEFT(REPLACE({tsv},'T',' '),19) || '.' || "
            f"RPAD(LEFT(SUBSTRING({tsv} FROM POSITION('.' IN {tsv})+1), {p}), {p}, '0')"
        )

    return f"""CASE
  /* NULL / null-like (incl. Hive's \\N export marker) */
  WHEN {c} IS NULL THEN NULL
  WHEN {ltv} IN ('','null','none','nan','nat','n/a','na','#n/a','<null>') THEN NULL
  WHEN {tv} = CHR(92)||'N' THEN NULL

  /* Booleans */
  WHEN {ltv} IN ('true','yes','y','1') THEN 'true'
  WHEN {ltv} IN ('false','no','n','0') THEN 'false'

  /* Timestamp WITH fractional seconds (on tz-stripped candidate) */
  WHEN LENGTH({tsv}) >= 21
       AND ({tsv} LIKE '____-__-__ __:__:__.%' OR {tsv} LIKE '____-__-__T__:__:__.%')
       AND TRY_CAST(REPLACE(LEFT({tsv},19),'T',' ') AS TIMESTAMP) IS NOT NULL
    THEN {ts_frac_result}

  /* Timestamp WITHOUT fractional seconds (on tz-stripped candidate) */
  WHEN LENGTH({tsv}) = 19
       AND ({tsv} LIKE '____-__-__ __:__:__' OR {tsv} LIKE '____-__-__T__:__:__')
       AND TRY_CAST(REPLACE({tsv},'T',' ') AS TIMESTAMP) IS NOT NULL
    THEN REPLACE({tsv},'T',' ')

  /* Scientific notation (Hive doubles: 1.5E7) → expand to positional via
     exact string→DECIMAL(33,18) cast (always renders positionally), then
     truncate to {prec} and strip zeros (truncation, never rounding). */
  WHEN CONTAINS(LOWER({cn}),'e')
       AND TRY_CAST({cn} AS DECIMAL(33,18)) IS NOT NULL
    THEN
      CASE
        WHEN RTRIM(LEFT(SPLIT_PART(CAST(TRY_CAST({cn} AS DECIMAL(33,18)) AS VARCHAR),'.',2),{prec}),'0') = ''
          THEN SPLIT_PART(CAST(TRY_CAST({cn} AS DECIMAL(33,18)) AS VARCHAR),'.',1)
        ELSE SPLIT_PART(CAST(TRY_CAST({cn} AS DECIMAL(33,18)) AS VARCHAR),'.',1)
             || '.' ||
             RTRIM(LEFT(SPLIT_PART(CAST(TRY_CAST({cn} AS DECIMAL(33,18)) AS VARCHAR),'.',2),{prec}),'0')
      END

  /* Pure integer (no dot, no 'e', castable to HUGEINT) */
  WHEN NOT CONTAINS({cn},'.')
       AND NOT CONTAINS(LOWER({cn}),'e')
       AND TRY_CAST({cn} AS HUGEINT) IS NOT NULL
    THEN CAST(CAST({cn} AS HUGEINT) AS VARCHAR)

  /* Whole-number decimal: 42.0, +42.0, -0.0 → canonical integer string.
     HUGEINT cast handles sign normalisation (+42→42, -0→0). */
  WHEN CONTAINS({cn},'.')
       AND NOT CONTAINS(LOWER({cn}),'e')
       AND TRY_CAST({cn} AS DOUBLE) IS NOT NULL
       AND RTRIM(SPLIT_PART({cn},'.',2),'0') = ''
       AND TRY_CAST(SPLIT_PART({cn},'.',1) AS HUGEINT) IS NOT NULL
    THEN CAST(TRY_CAST(SPLIT_PART({cn},'.',1) AS HUGEINT) AS VARCHAR)

  /* Decimal: truncate to {prec} places, strip trailing zeros */
  WHEN CONTAINS({cn},'.')
       AND NOT CONTAINS(LOWER({cn}),'e')
       AND TRY_CAST({cn} AS DOUBLE) IS NOT NULL
    THEN
      CASE
        WHEN LENGTH(SPLIT_PART({cn},'.',2)) > {prec} THEN
          CASE WHEN RTRIM(LEFT(SPLIT_PART({cn},'.',2),{prec}),'0') = ''
            THEN SPLIT_PART({cn},'.',1)
            ELSE SPLIT_PART({cn},'.',1) || '.' || RTRIM(LEFT(SPLIT_PART({cn},'.',2),{prec}),'0')
          END
        ELSE
          CASE WHEN RTRIM(SPLIT_PART({cn},'.',2),'0') = ''
            THEN SPLIT_PART({cn},'.',1)
            ELSE SPLIT_PART({cn},'.',1) || '.' || RTRIM(SPLIT_PART({cn},'.',2),'0')
          END
      END

  /* Everything else: return trimmed value */
  ELSE {tv}
END"""


# =============================================================================
# DUCKDB CSV LOADING
# =============================================================================

def load_csv_to_duckdb(con, filepath: str, table_name: str, source_name: str,
                       escape_char: Optional[str] = None) -> Tuple[List[str], int]:
    """Load CSV into a DuckDB table with preprocessing and delimiter detection."""
    with cleaned_csv(filepath) as (cleaned_filepath, rows_fixed):
        if rows_fixed > 0:
            logger.info(f"  [!] Fixed {rows_fixed:,} rows with errant surrounding quotes")

        hd, rd = detect_delimiters(cleaned_filepath)
        dn = {',': 'comma', '|': 'pipe', '\t': 'tab', '~': 'tilde', ';': 'semicolon'}
        if hd == rd:
            logger.info(f"  Detected delimiter: {dn.get(hd, repr(hd))}")
        else:
            logger.info(f"  Detected header delimiter: {dn.get(hd, repr(hd))}")
            logger.info(f"  Detected row delimiter: {dn.get(rd, repr(rd))}")

        con.execute(f"DROP TABLE IF EXISTS {table_name}")

        if hd != rd:
            # Mixed delimiter: load via pandas, register into DuckDB
            df = _load_mixed_delim(cleaned_filepath, hd, rd)
            df.columns = df.columns.str.strip().str.upper()
            con.register(f'_tmp_{table_name}', df)
            con.execute(f"CREATE TABLE {table_name} AS SELECT * FROM _tmp_{table_name}")
            con.unregister(f'_tmp_{table_name}')
        else:
            esc = f", escape='{escape_char}'" if escape_char else ""
            de = hd.replace("'", "''")
            if de == '\t':
                de = '\\t'
            try:
                con.execute(f"""CREATE TABLE {table_name} AS
                    SELECT * FROM read_csv('{cleaned_filepath}',
                        delim='{de}', header=true, all_varchar=true,
                        ignore_errors=true, null_padding=true{esc})""")
            except duckdb.Error as e:
                logger.warning(f"  [!] CSV read failed ({e}), trying auto-detect...")
                con.execute(f"""CREATE TABLE {table_name} AS
                    SELECT * FROM read_csv_auto('{cleaned_filepath}',
                        all_varchar=true, ignore_errors=true)""")

        # Uppercase column names (and strip UTF-8 BOM that export tools prepend
        # to the first column, which would otherwise break column alignment)
        cols = [r[0] for r in con.execute(
            f"SELECT column_name FROM information_schema.columns "
            f"WHERE table_name='{table_name}' ORDER BY ordinal_position").fetchall()]
        for col in cols:
            uc = col.lstrip('\ufeff').strip().upper()
            if uc != col:
                try:
                    con.execute(f'ALTER TABLE {table_name} RENAME COLUMN "{col}" TO "{uc}"')
                except duckdb.Error:
                    pass

        cols = [r[0] for r in con.execute(
            f"SELECT column_name FROM information_schema.columns "
            f"WHERE table_name='{table_name}' ORDER BY ordinal_position").fetchall()]

        # Handle single-column ROW_TEXT
        if len(cols) == 1 and cols[0] in ['ROW_TEXT', 'ROW', 'TEXT', 'DATA', 'LINE']:
            logger.info(f"  Detected single-column format: {cols[0]}")
            _expand_row_text(con, table_name, cols[0])
            cols = [r[0] for r in con.execute(
                f"SELECT column_name FROM information_schema.columns "
                f"WHERE table_name='{table_name}' ORDER BY ordinal_position").fetchall()]
            logger.info(f"  Expanded to {len(cols)} columns")

        rc = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
        logger.info(f"  [OK] Loaded {source_name}: {rc:,} rows, {len(cols)} columns")
        return cols, rc


def _load_mixed_delim(fp: str, hd: str, rd: str) -> pd.DataFrame:
    with open(fp, 'r', encoding='utf-8', errors='ignore') as f:
        hdr = [c.strip().upper() for c in next(csv.reader([f.readline().strip()], delimiter=hd))]
        rows = [next(csv.reader([l.strip()], delimiter=rd)) for l in f if l.strip()]
    df = pd.DataFrame(rows)
    if len(df.columns) == len(hdr):
        df.columns = hdr
    elif len(df.columns) < len(hdr):
        df.columns = hdr[:len(df.columns)]
    else:
        df.columns = hdr + [f'COL_{i+1}' for i in range(len(hdr), len(df.columns))]
    return df


def _expand_row_text(con, table_name: str, col_name: str):
    sample = con.execute(f'SELECT "{col_name}" FROM {table_name} WHERE "{col_name}" IS NOT NULL LIMIT 10').fetchall()
    if not sample:
        return
    vals = [r[0] for r in sample if r[0]]
    ds = {'|': 0, ',': 0, '\t': 0, ';': 0, '~': 0}
    for v in vals:
        for d in ds:
            ds[d] += str(v).count(d)
    delim = max(ds, key=ds.get)
    if ds[delim] == 0:
        return
    fv = next((v for v in vals if delim in str(v)), None)
    if not fv:
        return
    nc = len(fv.split(delim))
    exprs = [f"SPLIT_PART(\"{col_name}\",'{delim}',{i+1}) AS \"COL_{i+1}\"" for i in range(nc)]
    con.execute(f"CREATE TABLE _exp AS SELECT {','.join(exprs)} FROM {table_name}")
    con.execute(f"DROP TABLE {table_name}")
    con.execute(f"ALTER TABLE _exp RENAME TO {table_name}")


# =============================================================================
# DUCKDB CONNECTION HELPER
# =============================================================================

def create_duckdb_connection() -> duckdb.DuckDBPyConnection:
    """
    Create a DuckDB connection configured for large CSV comparisons.

    Uses an in-memory database with temporary directory spilling enabled
    so that intermediate results that exceed RAM are written to disk
    instead of causing an OOM crash.
    """
    con = duckdb.connect()
    # Allow DuckDB to spill to disk when memory is tight. Use the platform
    # temp directory (so it works on Windows, not just Unix /tmp) and create
    # it up front so spilling doesn't fail on a missing path under load.
    try:
        spill_dir = os.path.join(tempfile.gettempdir(), 'duckdb_temp')
        os.makedirs(spill_dir, exist_ok=True)
        con.execute(f"SET temp_directory='{spill_dir}'")
    except (duckdb.Error, OSError):
        pass  # older DuckDB versions may not support this
    # Log memory settings
    try:
        mem = con.execute("SELECT current_setting('memory_limit')").fetchone()[0]
        threads = con.execute("SELECT current_setting('threads')").fetchone()[0]
        logger.info(f"  DuckDB: memory_limit={mem}, threads={threads}")
    except duckdb.Error:
        pass
    return con


# =============================================================================
# DUCKDB COMPARATOR
# =============================================================================

class DuckDBComparator:
    """High-performance CSV comparison engine using DuckDB."""

    def __init__(self, con, source_table: str, target_table: str,
                 source_columns: List[str], target_columns: List[str],
                 key_columns: Optional[List[str]] = None,
                 skip_normalisation: bool = False,
                 decimal_precision: int = DEFAULT_NUMERIC_PRECISION,
                 detect_duplicates: bool = False) -> None:
        self.con = con
        self.source_table = source_table
        self.target_table = target_table
        self.all_source_columns = source_columns
        self.all_target_columns = target_columns
        self.key_columns = key_columns
        self.skip_normalisation = skip_normalisation
        self.decimal_precision = decimal_precision
        self.detect_duplicates = detect_duplicates
        self.discrepancies: List[Dict[str, Any]] = []
        self.common_columns: List[str] = []

    def _q(self, col: str) -> str:
        """Quote a column name for SQL."""
        return '"' + col.replace('"', '""') + '"'

    def _align_columns(self) -> None:
        ss, ts = set(self.all_source_columns), set(self.all_target_columns)
        cs = ss & ts
        self.common_columns = [c for c in self.all_source_columns if c in cs]
        self.source_only_columns = [c for c in self.all_source_columns if c not in cs]
        self.target_only_columns = [c for c in self.all_target_columns if c not in cs]

        for col in self.source_only_columns:
            self.discrepancies.append({'discrepancy_type': 'MISSING_COLUMN', 'composite_key': 'N/A',
                'column_name': col, 'source_value': 'COLUMN_EXISTS', 'target_value': 'COLUMN_NOT_FOUND',
                'full_source_row': 'N/A', 'full_target_row': 'N/A'})
        for col in self.target_only_columns:
            self.discrepancies.append({'discrepancy_type': 'MISSING_COLUMN', 'composite_key': 'N/A',
                'column_name': col, 'source_value': 'COLUMN_NOT_FOUND', 'target_value': 'COLUMN_EXISTS',
                'full_source_row': 'N/A', 'full_target_row': 'N/A'})

        if self.source_only_columns:
            logger.info(f"  [!] Columns only in source ({len(self.source_only_columns)}): {', '.join(self.source_only_columns)}")
        if self.target_only_columns:
            logger.info(f"  [!] Columns only in target ({len(self.target_only_columns)}): {', '.join(self.target_only_columns)}")
        logger.info(f"  [!] Missing columns excluded from hash and deep comparison")
        logger.info(f"  Common columns for comparison: {len(self.common_columns)}")

    def _create_normalised_tables(self) -> None:
        """No longer used — normalisation is done in Python on unmatched rows only."""
        pass

    def _compute_raw_hashes(self) -> None:
        """
        Compute MD5 hash of raw (un-normalised) column values using SQL.

        With 671 columns, the CONCAT_WS expression is ~40 KB — well within
        DuckDB's parser limits because there are no CASE expressions.  This
        catches all byte-identical rows in one fast SQL pass.

        When skip_normalisation is True, hashing is case-sensitive so that
        TRUE vs true is detected as a difference.
        """
        HASH_BATCH = 100  # columns per CONCAT_WS call
        # Case-insensitive by default; case-sensitive in skip_normalisation mode
        if self.skip_normalisation:
            coalesce_tpl = "COALESCE(CAST({col} AS VARCHAR),'\\x01NULL')"
        else:
            coalesce_tpl = "COALESCE(LOWER(CAST({col} AS VARCHAR)),'\\x01NULL')"

        def _build(table_in: str, table_out: str, label: str) -> None:
            logger.info(f"  Computing raw hashes for {label}...")
            self.con.execute(f"DROP TABLE IF EXISTS {table_out}")

            if len(self.common_columns) <= HASH_BATCH:
                parts = [coalesce_tpl.format(col=self._q(c)) for c in self.common_columns]
                hexpr = f"MD5(CONCAT_WS(CHR(0),{','.join(parts)}))"
                self.con.execute(f"CREATE TABLE {table_out} AS SELECT rowid AS _rid_, {hexpr} AS _h_ FROM {table_in}")
            else:
                partials = []
                for bi in range(0, len(self.common_columns), HASH_BATCH):
                    batch = self.common_columns[bi:bi + HASH_BATCH]
                    parts = [coalesce_tpl.format(col=self._q(c)) for c in batch]
                    partials.append(f"CONCAT_WS(CHR(0),{','.join(parts)})")
                hexpr = f"MD5(CONCAT_WS(CHR(0),{','.join(partials)}))"
                logger.info(f"    Wide table: hashing {len(self.common_columns)} cols in {len(partials)} batches...")
                self.con.execute(f"CREATE TABLE {table_out} AS SELECT rowid AS _rid_, {hexpr} AS _h_ FROM {table_in}")

        _build(self.source_table, "src_h", "source")
        _build(self.target_table, "tgt_h", "target")
        logger.info(f"  Raw hashes computed")

    def _hash_match(self) -> Tuple[int, int, int]:
        """Match rows by hash. Returns (exact_matches, src_unmatched, tgt_unmatched)."""
        logger.info(f"  Hash-based matching...")

        self.con.execute("""CREATE OR REPLACE TABLE hc AS
            SELECT h, s.cnt AS sc, t.cnt AS tc FROM
            (SELECT _h_ AS h, COUNT(*) AS cnt FROM src_h GROUP BY _h_) s
            JOIN (SELECT _h_ AS h, COUNT(*) AS cnt FROM tgt_h GROUP BY _h_) t USING(h)""")

        for rh, sc, tc in self.con.execute("SELECT h,sc,tc FROM hc WHERE sc!=tc").fetchall():
            self.discrepancies.append({'discrepancy_type': 'DUPLICATE_COUNT_MISMATCH',
                'composite_key': f'HASH:{rh[:16]}', 'column_name': 'ALL',
                'source_value': f'{sc} occurrences', 'target_value': f'{tc} occurrences',
                'full_source_row': 'N/A', 'full_target_row': 'N/A'})

        self.con.execute("""CREATE OR REPLACE TABLE src_matched AS
            SELECT _rid_ FROM (
                SELECT _rid_, _h_, ROW_NUMBER() OVER (PARTITION BY _h_ ORDER BY _rid_) AS rn
                FROM src_h WHERE _h_ IN (SELECT h FROM hc)
            ) r JOIN hc ON r._h_=hc.h WHERE r.rn <= LEAST(hc.sc,hc.tc)""")

        self.con.execute("""CREATE OR REPLACE TABLE tgt_matched AS
            SELECT _rid_ FROM (
                SELECT _rid_, _h_, ROW_NUMBER() OVER (PARTITION BY _h_ ORDER BY _rid_) AS rn
                FROM tgt_h WHERE _h_ IN (SELECT h FROM hc)
            ) r JOIN hc ON r._h_=hc.h WHERE r.rn <= LEAST(hc.sc,hc.tc)""")

        self.con.execute("CREATE OR REPLACE TABLE src_um AS SELECT _rid_ FROM src_h WHERE _rid_ NOT IN (SELECT _rid_ FROM src_matched)")
        self.con.execute("CREATE OR REPLACE TABLE tgt_um AS SELECT _rid_ FROM tgt_h WHERE _rid_ NOT IN (SELECT _rid_ FROM tgt_matched)")

        em = self.con.execute("SELECT COUNT(*) FROM src_matched").fetchone()[0]
        su = self.con.execute("SELECT COUNT(*) FROM src_um").fetchone()[0]
        tu = self.con.execute("SELECT COUNT(*) FROM tgt_um").fetchone()[0]

        logger.info(f"  [OK] Exact matches (raw): {em:,} rows")
        logger.info(f"  Rows to normalise & compare: {su:,} source, {tu:,} target")
        return em, su, tu

    def _normalised_hash_match(self) -> Tuple[int, int, int]:
        """
        Normalise unmatched rows and hash-match, all inside DuckDB SQL.

        With 671 columns, a single normalisation CASE expression is ~500 chars,
        so doing all columns in one SQL statement would produce ~335 KB of SQL.
        To stay safe, columns are processed in batches of 30 (~15 KB per batch).

        Strategy:
          1. For each batch of ~30 columns, compute a partial normalised string
             via CONCAT_WS + CASE expressions → stored as a temp table (_partial_N)
          2. Join all partials on _rid_ and compute MD5 of the combined string
          3. Match source vs target by normalised hash in SQL
          4. Update src_um / tgt_um tables with only the remaining unmatched rows

        No data leaves DuckDB — no pandas, no OOM.
        """
        NORM_BATCH = 30  # columns per SQL batch

        def _build_norm_hash_table(raw_table: str, um_table: str, hash_table: str, label: str) -> None:
            """Build a normalised hash table for one side (source or target)."""
            num_batches = (len(self.common_columns) + NORM_BATCH - 1) // NORM_BATCH
            logger.info(f"  Normalising {label}: {len(self.common_columns)} cols in {num_batches} batches...")

            partial_tables = []
            for bi in range(0, len(self.common_columns), NORM_BATCH):
                batch = self.common_columns[bi:bi + NORM_BATCH]
                batch_idx = bi // NORM_BATCH
                ptable = f"_partial_{label}_{batch_idx}"
                partial_tables.append(ptable)

                # Build normalisation expressions for this batch
                norm_parts = []
                for col in batch:
                    col_expr = self._q(col)
                    norm_sql = _build_normalise_sql(
                        col_expr, self.skip_normalisation,
                        self.decimal_precision, DEFAULT_TIMESTAMP_PRECISION
                    )
                    # Wrap: COALESCE([LOWER](normalised), chr(1)||'NULL')
                    # Case-sensitive when skip_normalisation is True
                    if self.skip_normalisation:
                        norm_parts.append(
                            f"COALESCE(CAST(({norm_sql}) AS VARCHAR), CHR(1)||'NULL')"
                        )
                    else:
                        norm_parts.append(
                            f"COALESCE(LOWER(CAST(({norm_sql}) AS VARCHAR)), CHR(1)||'NULL')"
                        )

                concat_expr = f"CONCAT_WS(CHR(0), {', '.join(norm_parts)})"

                self.con.execute(f"DROP TABLE IF EXISTS {ptable}")
                self.con.execute(f"""
                    CREATE TABLE {ptable} AS
                    SELECT rowid AS _rid_, {concat_expr} AS _p_
                    FROM {raw_table}
                    WHERE rowid IN (SELECT _rid_ FROM {um_table})
                """)

                if (batch_idx + 1) % 5 == 0 or batch_idx == num_batches - 1:
                    logger.info(f"    {label} batch {batch_idx+1}/{num_batches} done")

            # Join all partials and compute final MD5 hash
            if len(partial_tables) == 1:
                combined = f"SELECT _rid_, MD5(_p_) AS _nh_ FROM {partial_tables[0]}"
            else:
                # Join all partial tables on _rid_
                join_clauses = []
                concat_parts = []
                base = partial_tables[0]
                concat_parts.append(f"{base}._p_")
                for pt in partial_tables[1:]:
                    join_clauses.append(f"JOIN {pt} ON {base}._rid_ = {pt}._rid_")
                    concat_parts.append(f"{pt}._p_")

                full_concat = f"CONCAT_WS(CHR(0), {', '.join(concat_parts)})"
                combined = f"""
                    SELECT {base}._rid_, MD5({full_concat}) AS _nh_
                    FROM {base}
                    {' '.join(join_clauses)}
                """

            self.con.execute(f"DROP TABLE IF EXISTS {hash_table}")
            self.con.execute(f"CREATE TABLE {hash_table} AS {combined}")

            # Clean up partial tables
            for pt in partial_tables:
                self.con.execute(f"DROP TABLE IF EXISTS {pt}")

            logger.info(f"  {label} normalised hashes computed")

        # Build normalised hash tables for both sides
        _build_norm_hash_table(self.source_table, "src_um", "src_nh", "source")
        _build_norm_hash_table(self.target_table, "tgt_um", "tgt_nh", "target")

        # Match by normalised hash (same logic as raw hash match)
        self.con.execute("""CREATE OR REPLACE TABLE nhc AS
            SELECT h, s.cnt AS sc, t.cnt AS tc FROM
            (SELECT _nh_ AS h, COUNT(*) AS cnt FROM src_nh GROUP BY _nh_) s
            JOIN (SELECT _nh_ AS h, COUNT(*) AS cnt FROM tgt_nh GROUP BY _nh_) t USING(h)""")

        # DUPLICATE_COUNT_MISMATCH
        for rh, sc, tc in self.con.execute("SELECT h,sc,tc FROM nhc WHERE sc!=tc").fetchall():
            self.discrepancies.append({'discrepancy_type': 'DUPLICATE_COUNT_MISMATCH',
                'composite_key': f'HASH:{rh[:16]}', 'column_name': 'ALL',
                'source_value': f'{sc} occurrences', 'target_value': f'{tc} occurrences',
                'full_source_row': 'N/A', 'full_target_row': 'N/A'})

        # Pair matched rows 1-to-1
        self.con.execute("""CREATE OR REPLACE TABLE src_nm AS
            SELECT _rid_ FROM (
                SELECT _rid_, _nh_, ROW_NUMBER() OVER (PARTITION BY _nh_ ORDER BY _rid_) AS rn
                FROM src_nh WHERE _nh_ IN (SELECT h FROM nhc)
            ) r JOIN nhc ON r._nh_=nhc.h WHERE r.rn <= LEAST(nhc.sc, nhc.tc)""")

        self.con.execute("""CREATE OR REPLACE TABLE tgt_nm AS
            SELECT _rid_ FROM (
                SELECT _rid_, _nh_, ROW_NUMBER() OVER (PARTITION BY _nh_ ORDER BY _rid_) AS rn
                FROM tgt_nh WHERE _nh_ IN (SELECT h FROM nhc)
            ) r JOIN nhc ON r._nh_=nhc.h WHERE r.rn <= LEAST(nhc.sc, nhc.tc)""")

        norm_matches = self.con.execute("SELECT COUNT(*) FROM src_nm").fetchone()[0]

        # Update src_um / tgt_um: remove the rows that matched in this step
        self.con.execute("CREATE OR REPLACE TABLE src_um AS SELECT _rid_ FROM src_um WHERE _rid_ NOT IN (SELECT _rid_ FROM src_nm)")
        self.con.execute("CREATE OR REPLACE TABLE tgt_um AS SELECT _rid_ FROM tgt_um WHERE _rid_ NOT IN (SELECT _rid_ FROM tgt_nm)")

        su = self.con.execute("SELECT COUNT(*) FROM src_um").fetchone()[0]
        tu = self.con.execute("SELECT COUNT(*) FROM tgt_um").fetchone()[0]

        logger.info(f"  [OK] Normalised matches: {norm_matches:,} additional rows")
        logger.info(f"  Rows for deep comparison: {su:,} source, {tu:,} target")

        # Clean up
        for t in ['src_nh', 'tgt_nh', 'nhc', 'src_nm', 'tgt_nm']:
            self.con.execute(f"DROP TABLE IF EXISTS {t}")

        return norm_matches, su, tu

    def _compute_row_hashes(self) -> None:
        """Alias for backwards compatibility."""
        self._compute_raw_hashes()

    def _detect_composite_key(self) -> List[str]:
        logger.info("  Auto-detecting composite key columns...")
        total = self.con.execute("SELECT COUNT(*) FROM src_um").fetchone()[0]
        if total == 0:
            return self.common_columns[:20]
        logger.info(f"  Total columns: {len(self.common_columns)}")

        # Sample for type analysis
        sz = min(1000, total)
        scols = ','.join([self._q(c) for c in self.common_columns])
        sdf = self.con.execute(f"SELECT {scols} FROM {self.source_table} WHERE rowid IN (SELECT _rid_ FROM src_um) USING SAMPLE {sz} ROWS").fetchdf()
        sdf.columns = [c.upper() for c in sdf.columns]

        col_stats = {}
        excluded = []
        for col in self.common_columns:
            if col.startswith('_'):
                continue

            # Check if column matches a known key candidate pattern
            is_key_candidate = any(re.match(p, col, re.IGNORECASE) for p in KEY_CANDIDATE_PATTERNS)

            # Exclusion check — key candidates override name-based exclusion
            if not is_key_candidate and any(re.match(p, col, re.IGNORECASE) for p in COLUMN_EXCLUSION_PATTERNS):
                excluded.append((col, 'name pattern')); continue

            ct, tb = self._analyse_col(sdf[col] if col in sdf.columns else pd.Series(dtype=str))

            # Value-type exclusion — key candidates override timestamp/date exclusion
            # (columns like AS_OF_DATE contain dates but are valid composite key components)
            if ct in ('timestamp', 'filename', 'long_text', 'boolean'):
                if is_key_candidate:
                    # Override: keep the column but neutralise the negative type boost
                    tb = max(tb, 100)  # boost key candidate dates/timestamps
                    logger.debug(f"  Key candidate override: {col} ({ct} values kept as key candidate)")
                else:
                    excluded.append((col, f'{ct} values')); continue
            if ct == 'empty':
                excluded.append((col, 'all null-like')); continue

            # Name-based priority boost for key candidates
            if is_key_candidate:
                tb = max(tb, 120)  # ensure key candidates rank highly

            uc, nnc = self.con.execute(f"""SELECT
                COUNT(DISTINCT CASE WHEN {self._q(col)} IS NOT NULL
                  AND LOWER(TRIM(CAST({self._q(col)} AS VARCHAR))) NOT IN ('','null','none','nan','nat','n/a','na','#n/a','<null>')
                  THEN {self._q(col)} END),
                COUNT(CASE WHEN {self._q(col)} IS NOT NULL
                  AND LOWER(TRIM(CAST({self._q(col)} AS VARCHAR))) NOT IN ('','null','none','nan','nat','n/a','na','#n/a','<null>')
                  THEN 1 END)
                FROM {self.source_table} WHERE rowid IN (SELECT _rid_ FROM src_um)""").fetchone()

            nnr = nnc / total if total else 0
            ur = uc / total if total else 0
            pri = tb + ur * 50 + nnr * 30
            if uc < 10: pri -= 80
            if nnr < 0.5: pri -= 50
            col_stats[col] = {'priority': pri, 'unique_ratio': ur, 'non_null_ratio': nnr,
                              'unique_count': uc, 'col_type': ct, 'key_candidate': is_key_candidate}

        if excluded:
            logger.info(f"  Excluded {len(excluded)} column(s) from key detection: {', '.join(f'{c}({r})' for c,r in excluded[:5])}"
                        + (f" and {len(excluded)-5} more" if len(excluded) > 5 else ""))

        # Log key candidate overrides
        boosted = [c for c, s in col_stats.items() if s.get('key_candidate')]
        if boosted:
            logger.info(f"  Key candidate columns (name-boosted): {', '.join(boosted)}")

        sc = sorted(col_stats.items(), key=lambda x: -x[1]['priority'])
        if sc:
            top_display = ', '.join(f"{c}({s['col_type']}{'*' if s.get('key_candidate') else ''})" for c, s in sc[:5])
            logger.info(f"  Top key candidates: {top_display}  (* = name-boosted)")

        cands = [c for c, _ in sc[:30]]

        # Single unique column?
        result = []
        for c, s in sc:
            if s['unique_ratio'] == 1.0 and s['non_null_ratio'] == 1.0:
                result = [c]; break

        # Try combinations
        if not result:
            for i in range(1, min(len(cands)+1, 25)):
                combo = cands[:i]
                ce = " || '||' || ".join([f"COALESCE(CAST({self._q(c)} AS VARCHAR),'NULL')" for c in combo])
                ur = self.con.execute(f"SELECT CAST(COUNT(DISTINCT({ce})) AS DOUBLE)/COUNT(*) FROM {self.source_table} WHERE rowid IN (SELECT _rid_ FROM src_um)").fetchone()[0]
                if ur >= 0.999:
                    result = combo; break

        if not result:
            result = cands[:20] if cands else self.common_columns[:20]

        ce = " || '||' || ".join([f"COALESCE(CAST({self._q(c)} AS VARCHAR),'NULL')" for c in result])
        ur = self.con.execute(f"SELECT CAST(COUNT(DISTINCT({ce})) AS DOUBLE)/COUNT(*) FROM {self.source_table} WHERE rowid IN (SELECT _rid_ FROM src_um)").fetchone()[0]
        logger.info(f"  Selected {len(result)} key column(s): {', '.join(result)}")
        logger.info(f"  Composite key uniqueness: {ur:.4%}")
        return result

    @staticmethod
    def _analyse_col(vals: pd.Series) -> Tuple[str, int]:
        nn = vals.dropna()
        if len(nn) == 0: return 'empty', -100
        nn = nn[~nn.apply(_is_null_like)]
        if len(nn) == 0: return 'empty', -100
        sz = min(50, len(nn))
        smp = nn.iloc[::max(1, len(nn)//sz)][:sz].astype(str) if sz < len(nn) else nn.astype(str)
        ints=decs=alnum=alpha=ts=fn=bools=txt=ltxt=0
        bv = {'true','false','yes','no','y','n','0','1','t','f'}
        tp = [r'^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}', r'^\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}', r'^\d{4}-\d{2}-\d{2}$']
        fp = [r'.*\.(csv|txt|dat|xml|json|xlsx?|parquet)$', r'^[A-Za-z0-9_-]+_\d{8}.*\.']
        for v in smp:
            s=str(v).strip(); lo=s.lower()
            if lo in bv: bools+=1; continue
            if any(re.match(p,s) for p in tp): ts+=1; continue
            if any(re.search(p,s,re.I) for p in fp): fn+=1; continue
            if re.match(r'^-?\d+$',s): ints+=1; continue
            if re.match(r'^-?\d+\.\d+$',s): decs+=1; continue
            if re.match(r'^[A-Za-z0-9_-]+$',s) and re.search(r'[A-Za-z]',s) and re.search(r'\d',s): alnum+=1; continue
            if re.match(r'^[A-Za-z_-]+$',s) and len(s)<=20: alpha+=1; continue
            if len(s.split())>=5 or len(s)>=50: ltxt+=1; continue
            txt+=1
        t=len(smp)
        if alnum/t>0.5: return 'alphanumeric',150
        if ints/t>0.5: return 'integer',140
        if alpha/t>0.5: return 'alphabetic_code',80
        if ts/t>0.3: return 'timestamp',-200
        if fn/t>0.3: return 'filename',-200
        if ltxt/t>0.3: return 'long_text',-200
        if bools/t>0.3: return 'boolean',-100
        if decs/t>0.5: return 'decimal',-50
        if txt/t>0.5: return 'text',-30
        return 'mixed',0

    def _detect_duplicates_sql(self, label: str, key_cols: List[str], compare_cols: List[str]):
        hp = [f"COALESCE(LOWER(CAST({self._q(c)} AS VARCHAR)),CHR(1)||'NULL')" for c in compare_cols]
        hexpr = f"MD5(CONCAT_WS(CHR(0),{','.join(hp)}))"
        kp = [f"'{c}='||COALESCE(CAST({self._q(c)} AS VARCHAR),'NULL')" for c in key_cols]
        kexpr = "||'||'||".join(kp)
        rp = [f"COALESCE(CAST({self._q(c)} AS VARCHAR),'NULL')" for c in compare_cols]
        rexpr = "||'|'||".join(rp)
        vn = self.source_table if label == 'SOURCE' else self.target_table

        dups = self.con.execute(f"""WITH h AS (
            SELECT {hexpr} AS fh, {kexpr} AS ck, {rexpr} AS fr,
                   COUNT(*) OVER (PARTITION BY {hexpr}) AS dc FROM {vn})
            SELECT DISTINCT ck, fr, dc FROM h WHERE dc > 1""").fetchall()

        if dups:
            logger.info(f"    {label}: Found {len(dups)} unique rows duplicated ({sum(r[2] for r in dups)} total)")
            dt = f'DUPLICATE_IN_{label}'
            for ck, fr, dc in dups:
                self.discrepancies.append({'discrepancy_type': dt, 'composite_key': ck,
                    'column_name': f'IDENTICAL_ROWS_COUNT={dc}',
                    'source_value': fr if label=='SOURCE' else 'N/A',
                    'target_value': fr if label=='TARGET' else 'N/A',
                    'full_source_row': fr if label=='SOURCE' else 'N/A',
                    'full_target_row': fr if label=='TARGET' else 'N/A'})
        else:
            logger.info(f"    {label}: No duplicate rows found")

    def _deep_comparison(self, key_cols: List[str], compare_cols: List[str]):
        """Pull unmatched rows into pandas for deep comparison with fuzzy matching."""
        logger.info("  Loading unmatched rows for deep comparison...")
        cc = ','.join([self._q(c) for c in self.common_columns])

        # Fetch raw unmatched rows (normalisation done in Python below)
        src_raw = self.con.execute(f"SELECT rowid AS _rid_,{cc} FROM {self.source_table} WHERE rowid IN (SELECT _rid_ FROM src_um)").fetchdf()
        src_raw.columns = ['_RID_'] + [c.upper() for c in src_raw.columns[1:]]
        tgt_raw = self.con.execute(f"SELECT rowid AS _rid_,{cc} FROM {self.target_table} WHERE rowid IN (SELECT _rid_ FROM tgt_um)").fetchdf()
        tgt_raw.columns = ['_RID_'] + [c.upper() for c in tgt_raw.columns[1:]]

        # Original rows (all columns) for full-row output
        soc = ','.join([self._q(c) for c in self.all_source_columns])
        toc = ','.join([self._q(c) for c in self.all_target_columns])
        src_o = self.con.execute(f"SELECT rowid AS _rid_,{soc} FROM {self.source_table} WHERE rowid IN (SELECT _rid_ FROM src_um)").fetchdf()
        src_o.columns = ['_RID_'] + [c.upper() for c in src_o.columns[1:]]
        tgt_o = self.con.execute(f"SELECT rowid AS _rid_,{toc} FROM {self.target_table} WHERE rowid IN (SELECT _rid_ FROM tgt_um)").fetchdf()
        tgt_o.columns = ['_RID_'] + [c.upper() for c in tgt_o.columns[1:]]

        logger.info(f"  Loaded {len(src_raw):,} source and {len(tgt_raw):,} target unmatched rows")
        if len(src_raw) == 0 and len(tgt_raw) == 0:
            return

        # Normalise unmatched rows in Python
        skip_norm = self.skip_normalisation
        dec_prec = self.decimal_precision
        logger.info(f"  Normalising unmatched rows in Python...")

        def normalise_row_dict(row, cols):
            return {c: normalise_value(row[c], skip_norm, numeric_precision=dec_prec) for c in cols}

        src_n = src_raw  # keep reference for iteration
        tgt_n = tgt_raw

        # Build target index
        logger.info("  Building target index...")
        tgt_dict = {}; tgt_orig = {}
        tki: DefaultDict[str, List[int]] = defaultdict(list)
        tfki: DefaultDict[str, List[int]] = defaultdict(list)

        for i in range(len(tgt_n)):
            r = tgt_n.iloc[i]; idx = int(r['_RID_'])
            rd = normalise_row_dict(r, self.common_columns)
            tgt_dict[idx] = rd
            ro = tgt_o.iloc[i]
            tgt_orig[idx] = {c: ro[c] for c in self.all_target_columns if c in ro.index}
            key = '||'.join([f"{c}={str(rd[c]) if rd[c] is not None and not pd.isna(rd[c]) else 'NULL'}" for c in key_cols])
            tki[key].append(idx)
            tfki[build_fuzzy_key(rd, key_cols)].append(idx)

        dk = {k: v for k, v in tki.items() if len(v) > 1}
        if dk:
            logger.warning(f"  [!] Warning: {len(dk)} duplicate keys in unmatched target rows")
        logger.info(f"  Built {len(tki):,} exact keys, {len(tfki):,} fuzzy keys")

        logger.info("  Running deep comparison (with fuzzy key matching)...")
        matched_tgt: Set[int] = set()

        for i in range(len(src_n)):
            sr = src_n.iloc[i]; sid = int(sr['_RID_'])
            srd = normalise_row_dict(sr, self.common_columns)
            so = src_o.iloc[i]
            sod = {c: so[c] for c in self.all_source_columns if c in so.index}

            sk = '||'.join([f"{c}={str(srd[c]) if srd[c] is not None and not pd.isna(srd[c]) else 'NULL'}" for c in key_cols])
            mid = None; is_fuzzy = False; fz_diffs: List[Dict] = []

            # Exact key match
            if sk in tki:
                cands = [x for x in tki[sk] if x not in matched_tgt]
                if len(cands) == 1:
                    mid = cands[0]
                elif len(cands) > 1:
                    best, bs = None, -1
                    for ti in cands:
                        s = self._cmatch(srd, tgt_dict[ti], compare_cols, key_cols)
                        if s > bs: bs, best = s, ti
                    mid = best

            # Fuzzy key match
            if mid is None:
                sfk = build_fuzzy_key(srd, key_cols)
                if sfk in tfki:
                    fcands = [x for x in tfki[sfk] if x not in matched_tgt]
                    bfi, bfs, bfd = None, -1, []
                    for ti in fcands:
                        tr = tgt_dict[ti]; ok = True; td: List[Dict] = []
                        for c in key_cols:
                            sim, mt = key_values_fuzzy_equal(srd[c], tr[c], self.skip_normalisation, self.decimal_precision)
                            if not sim: ok = False; break
                            if mt != 'EXACT':
                                sn = normalise_value(srd[c], self.skip_normalisation, numeric_precision=self.decimal_precision)
                                tn = normalise_value(tr[c], self.skip_normalisation, numeric_precision=self.decimal_precision)
                                td.append({'column': c, 'source': str(sn) if sn else 'NULL', 'target': str(tn) if tn else 'NULL'})
                        if ok:
                            s = self._cmatch(srd, tr, compare_cols, key_cols)
                            if s > bfs: bfs, bfi, bfd = s, ti, td
                    if bfi is not None:
                        mid = bfi; is_fuzzy = True; fz_diffs = bfd

            # Compare matched
            if mid is not None:
                matched_tgt.add(mid)
                tr = tgt_dict[mid]; to = tgt_orig[mid]
                diffs = []
                if is_fuzzy and fz_diffs:
                    for kd in fz_diffs:
                        diffs.append({'column': kd['column'], 'source': kd['source'], 'target': kd['target'], 'is_key': True})
                rkc = {d['column'] for d in fz_diffs} if fz_diffs else set()
                for c in compare_cols:
                    if c in rkc: continue
                    sv, tv = srd[c], tr[c]
                    sn = pd.isna(sv) or sv is None; tn = pd.isna(tv) or tv is None
                    if sn and tn: continue
                    if sn or tn:
                        eq = False
                    else:
                        eq = normalised_values_equal(sv, tv, self.skip_normalisation, self.decimal_precision)
                    if not eq:
                        diffs.append({'column': c, 'source': str(sv) if not sn else 'NULL',
                                      'target': str(tv) if not tn else 'NULL', 'is_key': c in key_cols})
                if diffs:
                    fsr = '|'.join([str(sod.get(c,'N/A')) if not pd.isna(sod.get(c,None)) else 'NULL' for c in self.all_source_columns])
                    ftr = '|'.join([str(to.get(c,'N/A')) if not pd.isna(to.get(c,None)) else 'NULL' for c in self.all_target_columns])
                    tk = '||'.join([f"{c}={str(tr[c]) if tr[c] is not None and not pd.isna(tr[c]) else 'NULL'}" for c in key_cols])
                    for d in diffs:
                        self.discrepancies.append({
                            'discrepancy_type': 'KEY_VALUE_MISMATCH' if d.get('is_key') else 'VALUE_MISMATCH',
                            'composite_key': f"{sk} ~> {tk}" if is_fuzzy else sk,
                            'column_name': d['column'], 'source_value': d['source'], 'target_value': d['target'],
                            'full_source_row': fsr, 'full_target_row': ftr})
                continue

            # Missing in target
            frd = '|'.join([str(sod.get(c,'N/A')) if not pd.isna(sod.get(c,None)) else 'NULL' for c in self.all_source_columns])
            self.discrepancies.append({'discrepancy_type': 'MISSING_IN_TARGET', 'composite_key': sk,
                'column_name': 'ALL', 'source_value': frd, 'target_value': 'ROW_NOT_FOUND',
                'full_source_row': frd, 'full_target_row': 'ROW_NOT_FOUND'})

            if (i+1) % 10000 == 0 or (i+1) == len(src_n):
                logger.info(f"    Progress: {i+1:,}/{len(src_n):,} ({(i+1)/len(src_n)*100:.0f}%)")

        # Missing in source
        umtgt = set(tgt_dict.keys()) - matched_tgt
        if umtgt:
            logger.info(f"  Identifying {len(umtgt):,} rows only in target...")
            for ti in umtgt:
                tr = tgt_dict[ti]; to = tgt_orig[ti]
                k = '||'.join([f"{c}={str(tr[c]) if tr[c] is not None and not pd.isna(tr[c]) else 'NULL'}" for c in key_cols])
                fr = '|'.join([str(to.get(c,'N/A')) if not pd.isna(to.get(c,None)) else 'NULL' for c in self.all_target_columns])
                self.discrepancies.append({'discrepancy_type': 'MISSING_IN_SOURCE', 'composite_key': k,
                    'column_name': 'ALL', 'source_value': 'ROW_NOT_FOUND', 'target_value': fr,
                    'full_source_row': 'ROW_NOT_FOUND', 'full_target_row': fr})

        mc = len(matched_tgt)
        logger.info(f"\n  Comparison Summary:")
        logger.info(f"    Rows matched by key: {mc:,}")
        logger.info(f"    Rows only in source: {len(src_n)-mc:,}")
        logger.info(f"    Rows only in target: {len(umtgt):,}")

    @staticmethod
    def _cmatch(s: dict, t: dict, cc: List[str], kc: List[str]) -> int:
        m = 0
        for c in cc:
            if c in kc: continue
            sv, tv = s.get(c), t.get(c)
            sn = pd.isna(sv) or sv is None; tn = pd.isna(tv) or tv is None
            if sn and tn: m += 1
            elif not sn and not tn and (str(sv)==str(tv) or str(sv).lower()==str(tv).lower()): m += 1
        return m

    def compare(self) -> List[Dict[str, Any]]:
        logger.info("\n" + "=" * 60)
        logger.info("COMPREHENSIVE COMPARISON (DuckDB Engine)")
        logger.info("=" * 60)
        sys.stdout.flush()

        import time

        logger.info("\nStep 1: Aligning columns...")
        self._align_columns()
        sys.stdout.flush()

        t0 = time.time()
        logger.info("\nStep 2: Computing raw row hashes (SQL)...")
        try:
            self._compute_raw_hashes()
        except Exception as e:
            logger.error(f"\n[X] FATAL: Hash computation failed: {e}")
            raise CSVComparatorError(f"Hash computation failed: {e}") from e
        logger.info(f"  Step 2 completed in {time.time()-t0:.1f}s")
        sys.stdout.flush()

        t0 = time.time()
        logger.info("\nStep 3: Hash-based exact matching (SQL)...")
        raw_matches, su, tu = self._hash_match()
        logger.info(f"  Step 3 completed in {time.time()-t0:.1f}s")
        sys.stdout.flush()

        if su == 0 and tu == 0:
            if self.discrepancies:
                logger.info("\n[OK] All data rows matched on common columns (schema differences detected)")
            else:
                logger.info("\n[OK] All rows matched exactly!")
            return self.discrepancies

        t0 = time.time()
        logger.info("\nStep 4: Normalised hash matching (SQL)...")
        logger.info("  (catches format differences: 100.50 vs 100.5, NULL vs None, timestamps, booleans)")
        try:
            norm_matches, su2, tu2 = self._normalised_hash_match()
        except Exception as e:
            logger.error(f"\n[X] FATAL: Normalised matching failed: {e}")
            raise CSVComparatorError(f"Normalised matching failed: {e}") from e
        logger.info(f"  Step 4 completed in {time.time()-t0:.1f}s")
        sys.stdout.flush()

        if su2 == 0 and tu2 == 0:
            if self.discrepancies:
                logger.info("\n[OK] All data rows matched (schema differences or count mismatches detected)")
            else:
                logger.info("\n[OK] All rows matched after normalisation!")
            return self.discrepancies

        t0 = time.time()
        logger.info("\nStep 5: Deep key-based comparison...")
        if self.key_columns:
            kc = [c.upper() for c in self.key_columns if c.upper() in self.common_columns]
            logger.info(f"  Using specified key columns: {', '.join(kc)}")
        else:
            kc = self._detect_composite_key()

        cc = list(self.common_columns)
        if self.detect_duplicates:
            logger.info("\n  Detecting duplicate rows...")
            self._detect_duplicates_sql('SOURCE', kc, cc)
            self._detect_duplicates_sql('TARGET', kc, cc)
        else:
            logger.info("\n  Duplicate row detection: Skipped (use --detect-duplicates to enable)")

        self._deep_comparison(kc, cc)
        logger.info(f"  Step 5 completed in {time.time()-t0:.1f}s")
        sys.stdout.flush()
        return self.discrepancies


# =============================================================================
# REPORT GENERATION (unchanged from original)
# =============================================================================

def _write_xlsx_file_parallel(args):
    df, output_file, _, _ = args
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    bad = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ud800-\udfff\ufdd0-\ufddf\ufffe\uffff]')
    wb = Workbook(); ws = wb.active; ws.title = "Discrepancies"
    for ri, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=bad.sub('', val) if isinstance(val, str) else val)
            if ri == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
    for c, w in {'A':20,'B':50,'C':25,'D':40,'E':40,'F':60,'G':60}.items():
        ws.column_dimensions[c].width = w
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    wb.save(output_file)
    return output_file


def generate_report(discrepancies, output_file, column_headers=None, num_workers=None, output_format="xlsx"):
    logger.info("\n" + "=" * 60)
    logger.info("REPORT GENERATION")
    logger.info("=" * 60)
    if num_workers is None:
        num_workers = max(1, mp.cpu_count() - 1)
    base = os.path.splitext(output_file)[0]

    if not discrepancies:
        logger.info("\n[OK] SUCCESS: No discrepancies found!")
        df = pd.DataFrame(columns=['discrepancy_type','composite_key','column_name','source_value','target_value','full_source_row','full_target_row'])
        of = base + (".xlsx" if output_format == "xlsx" else ".csv")
        (df.to_excel if output_format == "xlsx" else df.to_csv)(of, index=False, **({"engine": "openpyxl"} if output_format == "xlsx" else {}))
        logger.info(f"\n[OK] Empty report saved to: {of}")
        return

    total = len(discrepancies)
    logger.info(f"\n  Processing {total:,} discrepancies...")
    df = pd.DataFrame(discrepancies)
    co = ['discrepancy_type','composite_key','column_name','source_value','target_value','full_source_row','full_target_row']
    df = df[[c for c in co if c in df.columns]]
    if total >= 10000:
        logger.info("  Sorting discrepancies...")
    df = df.sort_values(['discrepancy_type','composite_key','column_name'])

    if output_format == "xlsx":
        try:
            from openpyxl import Workbook
        except ImportError:
            logger.error("[X] openpyxl required for XLSX. pip install openpyxl"); raise
        MX = 1048575; np_ = (total + MX - 1) // MX; ofs = []
        if np_ == 1:
            logger.info("  Writing XLSX report...")
            of = base + ".xlsx"
            _write_xlsx_file_parallel((df, of, 1, 1)); ofs.append(of)
        else:
            logger.info(f"  Splitting into {np_} XLSX files...")
            args = [(df.iloc[p*MX:min((p+1)*MX,total)].copy(), f"{base}_part{p+1}.xlsx", p+1, np_) for p in range(np_)]
            with ProcessPoolExecutor(max_workers=min(num_workers, np_)) as ex:
                ofs = list(ex.map(_write_xlsx_file_parallel, args))
        output_file = ofs[0] if len(ofs)==1 else ofs
    else:
        output_file = base + ".csv"
        df.to_csv(output_file, index=False)

    logger.info(f"\n[X] DISCREPANCIES FOUND: {total:,}")
    logger.info("\nBreakdown by type:")
    for dt, cnt in df['discrepancy_type'].value_counts().items():
        logger.info(f"  {dt}: {cnt:,}")
    if 'VALUE_MISMATCH' in df['discrepancy_type'].values:
        logger.info(f"\nValue mismatches by column (top 10):")
        for c, n in df[df['discrepancy_type']=='VALUE_MISMATCH']['column_name'].value_counts().head(10).items():
            logger.info(f"  {c}: {n:,}")

    if isinstance(output_file, list):
        logger.info(f"\n[OK] Report saved to {len(output_file)} files:")
        for f in output_file: logger.info(f"  - {f}")
    else:
        logger.info(f"\n[OK] Report saved to: {output_file}")

    if column_headers:
        ref = base + "_column_reference.txt"
        with open(ref, 'w') as f:
            f.write("COLUMN REFERENCE FOR ROW DATA\n" + "="*60 + "\n\n")
            f.write("The full_source_row and full_target_row columns contain\npipe-delimited (|) data for the complete row.\n\n")
            if isinstance(column_headers, tuple) and len(column_headers) == 2:
                sc, tc = column_headers; ss, ts = set(sc), set(tc)
                f.write("SOURCE Row Column Order (full_source_row):\n")
                for i, c in enumerate(sc): f.write(f"  {i+1}. {c}{' [SOURCE ONLY]' if c not in ts else ''}\n")
                f.write(f"\nTotal source columns: {len(sc)}\n\n{'='*60}\n\n")
                f.write("TARGET Row Column Order (full_target_row):\n")
                for i, c in enumerate(tc): f.write(f"  {i+1}. {c}{' [TARGET ONLY]' if c not in ss else ''}\n")
                f.write(f"\nTotal target columns: {len(tc)}\n")
                so = [c for c in sc if c not in ts]; to = [c for c in tc if c not in ss]
                if so or to:
                    f.write(f"\n{'='*60}\n\nMISSING COLUMN SUMMARY\n")
                    if so: f.write(f"  Source-only ({len(so)}): {', '.join(so)}\n")
                    if to: f.write(f"  Target-only ({len(to)}): {', '.join(to)}\n")
        logger.info(f"[OK] Column reference saved to: {ref}")

    logger.info("\nSample discrepancies (first 5):")
    s = df.head(5).copy()
    for c in ['source_value','target_value']:
        s[c] = s[c].apply(lambda x: x[:50]+'...' if len(str(x))>50 else x)
    logger.info(s.to_string(index=False))


# =============================================================================
# BATCH MODE
# =============================================================================

def _extract_table_key(fn):
    bn = os.path.splitext(fn)[0]
    m = re.match(r'^(?:Source|Target)_(.+)$', bn, re.IGNORECASE)
    if not m: return None
    return re.sub(r'(_\d+)+$', '', m.group(1))


def match_csv_pairs(sd, td):
    sf, tf = {}, {}
    for fn in sorted(os.listdir(sd)):
        if fn.lower().endswith('.csv'):
            k = _extract_table_key(fn)
            if k: sf[k] = os.path.join(sd, fn)
    for fn in sorted(os.listdir(td)):
        if fn.lower().endswith('.csv'):
            k = _extract_table_key(fn)
            if k: tf[k] = os.path.join(td, fn)
    matched, us, ut = [], [], []
    for k in sorted(set(sf)|set(tf)):
        if k in sf and k in tf: matched.append((k, sf[k], tf[k]))
        elif k in sf: us.append(os.path.basename(sf[k]))
        else: ut.append(os.path.basename(tf[k]))
    return matched, us, ut


def run_single_comparison(table_key, source_file, target_file, output_dir,
                          skip_normalisation=False, decimal_precision=DEFAULT_NUMERIC_PRECISION,
                          escape_char=None, output_format="xlsx", detect_duplicates=False):
    lf = os.path.join(output_dir, f"{table_key}_comparison.log")
    fh = logging.FileHandler(lf, mode='w', encoding='utf-8')
    fh.setFormatter(logging.Formatter('%(message)s')); fh.setLevel(logging.DEBUG)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(logging.Formatter(f'[{table_key}] %(message)s')); ch.setLevel(logging.INFO)
    cl = logging.getLogger(f'csv_comparator.{table_key}')
    cl.setLevel(logging.DEBUG); cl.addHandler(fh); cl.addHandler(ch); cl.propagate = False
    global logger; ol = logger; logger = cl
    import csv_comparator_duckdb as _m; _m.logger = cl
    success, summary = False, "ERROR"
    try:
        cl.info("="*70); cl.info(f"CSV Comparator (DuckDB) - {table_key}"); cl.info("="*70)
        cl.info(f"\nSource: {source_file}\nTarget: {target_file}")
        cl.info(f"Normalisation: {'Off' if skip_normalisation else 'On'}, Precision: {decimal_precision}, Format: {output_format.upper()}, Duplicates: {'On' if detect_duplicates else 'Off'}")
        cl.info("\n"+"-"*40+"\nLoading files...")
        con = create_duckdb_connection()
        sc, _ = load_csv_to_duckdb(con, source_file, 'source_raw', 'Source', escape_char)
        tc, _ = load_csv_to_duckdb(con, target_file, 'target_raw', 'Target', escape_char)
        cmp = DuckDBComparator(con, 'source_raw', 'target_raw', sc, tc,
            skip_normalisation=skip_normalisation, decimal_precision=decimal_precision, detect_duplicates=detect_duplicates)
        discs = cmp.compare()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        generate_report(discs, os.path.join(output_dir, f"{table_key}_discrepancy_report_{ts}"),
            column_headers=(cmp.all_source_columns, cmp.all_target_columns), output_format=output_format)
        summary = "PASS - No discrepancies" if not discs else f"FAIL - {len(discs):,} discrepancies"
        success = True; con.close()
    except Exception as e:
        summary = f"ERROR - {e}"; cl.error(f"\n[X] Error: {e}", exc_info=True)
    finally:
        logger = ol; _m.logger = ol; cl.removeHandler(fh); cl.removeHandler(ch); fh.close()
    return table_key, success, summary


def run_batch_comparison(sd, td, args):
    logger.info(f"\nSource: {sd}\nTarget: {td}\nScanning...")
    matched, us, ut = match_csv_pairs(sd, td)
    if not matched and not us and not ut:
        logger.error("\n[X] No CSV files found."); sys.exit(1)
    logger.info(f"\n  Matched: {len(matched)}")
    if us: logger.info(f"  Unmatched source: {len(us)}")
    if ut: logger.info(f"  Unmatched target: {len(ut)}")
    if not matched:
        logger.error("\n[X] No matching pairs found."); sys.exit(1)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    od = args.output_dir or f"batch_comparison_{ts}"
    os.makedirs(od, exist_ok=True)
    logger.info(f"  Output: {od}\n{'='*70}\nStarting batch ({len(matched)} tables)...\n{'='*70}")
    results = []
    for i, (tk, sp, tp) in enumerate(matched, 1):
        logger.info(f"\n{'─'*70}\n  [{i}/{len(matched)}] {tk}\n{'─'*70}")
        results.append(run_single_comparison(tk, sp, tp, od, args.no_normalisation,
            args.decimal_precision, args.escape_char, args.output_format, args.detect_duplicates))
    # Summary
    ps = [r for r in results if "PASS" in r[2]]
    fl = [r for r in results if "FAIL" in r[2]]
    er = [r for r in results if "ERROR" in r[2]]
    sl = [""]+["="*70]+["BATCH COMPARISON SUMMARY"]+["="*70]
    sl += [f"  Total: {len(results)}, Passed: {len(ps)}, Failed: {len(fl)}, Errors: {len(er)}",""]
    ml = max((len(r[0]) for r in results), default=20)
    sl += [f"  {'Table':<{ml}}   Result", f"  {'─'*ml}   {'─'*30}"]
    for tk, _, sm in results: sl.append(f"  {tk:<{ml}}   {sm}")
    if us: sl += [f"\n  Unmatched source:"] + [f"    - {f}" for f in us]
    if ut: sl += [f"\n  Unmatched target:"] + [f"    - {f}" for f in ut]
    sl += ["", f"  Reports saved to: {od}", "="*70]
    for l in sl: logger.info(l)
    sf = os.path.join(od, f"batch_summary_{ts}.txt")
    with open(sf, 'w', encoding='utf-8') as f:
        for l in sl: f.write(l+'\n')
    logger.info(f"\n  Batch summary: {sf}")


# =============================================================================
# CLI
# =============================================================================

def parse_arguments():
    p = argparse.ArgumentParser(description="CSV Comparator (DuckDB) - Migration Validation",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Modes:\n  Single: python csv_comparator_duckdb.py source.csv target.csv [KEY_COL ...]\n  Batch:  python csv_comparator_duckdb.py ./source_csv ./target_csv")
    p.add_argument("source_csv", nargs="?")
    p.add_argument("target_csv", nargs="?")
    p.add_argument("key_columns", nargs="*")
    p.add_argument("--output-dir", dest="output_dir")
    p.add_argument("--no-normalisation", dest="no_normalisation", action="store_true", default=False)
    p.add_argument("--decimal-precision", dest="decimal_precision", type=int, default=6)
    p.add_argument("--esc-char", dest="escape_char", type=str, default=None)
    p.add_argument("--output-format", dest="output_format", type=str, default="xlsx", choices=["csv","xlsx"])
    p.add_argument("--detect-duplicates", dest="detect_duplicates", action="store_true", default=False)
    p.add_argument("-v", "--verbose", action="store_true")
    return p.parse_args()


def main():
    args = parse_arguments()
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    logger.info("="*70)
    logger.info("CSV Comparator (DuckDB Edition) - Migration Validation")
    logger.info("="*70)

    if args.source_csv and args.target_csv:
        if os.path.isdir(args.source_csv) and os.path.isdir(args.target_csv):
            run_batch_comparison(args.source_csv, args.target_csv, args); return
        sf, tf = args.source_csv, args.target_csv
        kc = args.key_columns or None
        od, sn, dp, ec, of_, dd = args.output_dir, args.no_normalisation, args.decimal_precision, args.escape_char, args.output_format, args.detect_duplicates
    else:
        logger.info("\nUsage: python csv_comparator_duckdb.py <source> <target> [keys...] [options]")
        sf = input("Source CSV: ").strip(); tf = input("Target CSV: ").strip()
        ki = input("Key columns (comma-separated or Enter): ").strip()
        kc = [c.strip() for c in ki.split(',')] if ki else None
        od, sn, dp, ec, of_, dd = None, False, DEFAULT_NUMERIC_PRECISION, None, "xlsx", False

    for path, name in [(sf,"Source"),(tf,"Target")]:
        if not os.path.exists(path):
            logger.error(f"\n[X] {name} not found: {path}"); sys.exit(1)

    logger.info(f"\nSource: {sf}\nTarget: {tf}\nKeys: {', '.join(kc) if kc else 'Auto-detect'}")
    if od: logger.info(f"Output dir: {od}")
    logger.info(f"Normalisation: {'Off' if sn else 'On'}, Precision: {dp}, Escape: {repr(ec) if ec else 'None'}")
    logger.info(f"Format: {of_.upper()}, Duplicates: {'On' if dd else 'Off'}, Engine: DuckDB {duckdb.__version__}")

    logger.info("\n"+"-"*40+"\nLoading files into DuckDB...")
    con = create_duckdb_connection()
    try:
        sc, _ = load_csv_to_duckdb(con, sf, 'source_raw', 'Source (Hive)', ec)
        tc, _ = load_csv_to_duckdb(con, tf, 'target_raw', 'Target (Snowflake)', ec)
    except CSVComparatorError:
        sys.exit(1)

    cmp = DuckDBComparator(con, 'source_raw', 'target_raw', sc, tc,
        key_columns=kc, skip_normalisation=sn, decimal_precision=dp, detect_duplicates=dd)
    discs = cmp.compare()

    bn = os.path.splitext(os.path.basename(sf))[0]
    pts = bn.rsplit('_', 2)
    tn = '_'.join(pts[:-2]) if len(pts)>=3 and pts[-2].isdigit() and pts[-1].isdigit() else bn
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    ofn = f"{tn}_discrepancy_report_{ts}"
    if od:
        os.makedirs(od, exist_ok=True); ofn = os.path.join(od, ofn)

    generate_report(discs, ofn, column_headers=(cmp.all_source_columns, cmp.all_target_columns), output_format=of_)
    logger.info("\n"+"="*70+"\nComparison Complete!\n"+"="*70)
    con.close()


if __name__ == "__main__":
    main()
